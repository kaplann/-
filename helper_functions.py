import pandas as pd
from openpyxl import load_workbook
import matplotlib.pylab as plt
import matplotlib
import numpy as np
from bokeh.plotting import figure, output_file, show
from bokeh.io import output_notebook
from sklearn.preprocessing import StandardScaler

import warnings
warnings.filterwarnings('ignore', category=DeprecationWarning, message='.*use @default decorator instead.*')
#from mpld3 import plugins
#import mpld3 
matplotlib.style.use('ggplot')
#%matplotlib inline
import bokeh
from bokeh.io import output_notebook
from bokeh.plotting import figure, show, output_file
from bokeh.models import ColumnDataSource, Range1d, LabelSet, Label, HoverTool
from bokeh.models.widgets import Panel, Tabs
from bokeh.charts import Histogram
from bokeh.layouts import gridplot

#########################  Bokeh  #######################################
from bokeh.io import output_notebook
from bokeh.plotting import figure, show, output_file
from bokeh.models import ColumnDataSource, Range1d, LabelSet, Label, HoverTool
from bokeh.models.widgets import Panel, Tabs
from bokeh.charts import Histogram
from bokeh.layouts import gridplot
import matplotlib as mpl

from matplotlib.mlab import bivariate_normal

from bokeh.plotting import figure, output_file, show
from bokeh.models import LogColorMapper, LogTicker, ColorBar
# 	output_notebook()

# Get number of NaN values in column: 
def count_NaNs(df, sheet, col): 
    return sum(1 for booli in df[sheet][col].isnull() if booli==True)

def count_NaN_symbol(df, sheet, col, symbols=['-', '..', '*', '**']): 
    return sum(1 for booli in df[sheet][col] if booli in symbols)

def create_sheet_translator(sheet_stats):
    new_sheets_orderd = [x for x in sheet_stats['Sheets']]
    
    pair_dict = {'Artzi': 'קובץ ארצי',
                 'Nafa': 'קובץ נפות',
                 'Mahoz': 'קובץ מחוזות',
                 'BigTown': 'קובץ יישובים גדולים (10,000 +)',
                 'SmallTown': ' יישובים בינוניים (2,000-9,999)',
                 'Councils': 'קובץ מועצות איזוריות'}
    return dict((key, get_index_sheet(pair_dict[key], new_sheets_orderd)) for key in pair_dict)

def get_index_sheet(key, sheets):
    return [num for num, sheet_name in enumerate(sheets) if sheet_name==key][0]

#from wand.image import Image as WImage
from IPython.core.display import HTML

from IPython.display import Image
#print(Image(filename='pictures/flyer_header_wide.png', width=1000, height=1000))


# Get set of intersecting columns from 2 sheets. 
def intersecting_feats(data, sheet1, sheet2): 
    return set(data[sheet1].keys()).intersection(set(data[sheet2].keys()))

# Get number of intersecting columns from 2 sheets. 
def num_intersecting_feats(data, sheet1, sheet2,): 
    return len(intersecting_feats(data, sheet1, sheet2))

# Get number of NaN values in column: 
def count_NaNs(df, sheet, col): 
    return sum(1 for booli in df[sheet][col].isnull() if booli==True)

# Translate Nafa name using number:
def nafa_name_translator(num): 
    num_name_dict = {21: 'Tzfat', 22: 'Kineret', 23: "Jezreel (Yizre'el)", 24: 'Akko', 29: 'Golan', 31: 'Haifa', 32: 'Hadera', 
                     41: 'Sharon', 42: 'Petach Tikva', 43: 'Ramla', 44: 'Rehovot', 61: 'Ashkelon', 62: "Be'er Sheva"}
    return num_name_dict[num]

# Return all columns with no NaNs from sheet: 
def cols_no_NaNs(data, sheet_name): 
    return [col for col in data[sheet_name].columns if count_NaNs(data, sheet_name, col)==0]

# Return number of columns with no NaN values. 
def num_cols_no_NaNs(data, sheet_name):
    return len(cols_no_NaNs(data, sheet_name))


# Return all columns with no NaNs or symbols from sheet: 
def cols_no_NaNs_symbol(data, sheet_name): 
    no_symbols = [col for col in data[sheet_name].columns if count_NaN_symbol(data, sheet_name, col)==0]
    no_Nans = [col for col in data[sheet_name].columns if count_NaNs(data, sheet_name, col)==0]
    return set(no_symbols).intersection(no_Nans)

# Return all columns with no NaNs or symbols from sheet: 
def cols_no_NaNs_symbol_parentheses(data, sheet_name): 
    no_symbols_par = [col for col in data[sheet_name].columns if count_NaN_symbol_parentheses(data, sheet_name, col)==0]
    no_Nans = [col for col in data[sheet_name].columns if count_NaNs(data, sheet_name, col)==0]
    return set(no_symbols_par).intersection(no_Nans)

# Return number of columns with no NaN or symbol values. 
def num_cols_no_NaNs_symbol(data, sheet_name):
    return len(cols_no_NaNs_symbol(data, sheet_name))

# Return number of columns with no NaN or symbol values. 
def num_cols_no_NaNs_symbol_par(data, sheet_name):
    return len(cols_no_NaNs_symbol_parentheses(data, sheet_name))


# Retrun number of symbol values in column: 
def count_NaN_symbol(df, sheet, col, symbols=['-', '..', '*', '**']): 
    return sum(1 for booli in df[sheet][col] if booli in symbols)

# Retrun number of symbol values in column: 
def count_NaN_symbol_parentheses(df, sheet, col, symbols=['-', '..', '*', '**']): 
    return sum(1 for booli in df[sheet][col] if (booli in symbols) or (type(booli)==str))

def create_sheet_translator(sheet_stats):
    new_sheets_orderd = [x for x in sheet_stats['Sheets']]
    
    pair_dict = {'Artzi': 'קובץ ארצי',
                 'Nafa': 'קובץ נפות',
                 'Mahoz': 'קובץ מחוזות',
                 'BigTown': 'קובץ יישובים גדולים (10,000 +)',
                 'SmallTown': ' יישובים בינוניים (2,000-9,999)',
                 'Councils': 'קובץ מועצות איזוריות'}
    return dict((key, get_index_sheet(pair_dict[key], new_sheets_orderd)) for key in pair_dict)

def get_index_sheet(key, sheets):
    return [num for num, sheet_name in enumerate(sheets) if sheet_name==key][0]

def calc_eig_vals(df_i):
    # Normalizing data: Substracting mean and dividing by variance. 
    X_std = StandardScaler().fit_transform(df_i)
    mean_vec = np.mean(X_std, axis=0)
    cov_mat = (X_std - mean_vec).T.dot((X_std - mean_vec)) / (X_std.shape[0]-1)
    eig_vals, eig_vecs = np.linalg.eig(cov_mat)
    return eig_vals, eig_vecs

def plot_scree(eig_vals, sheet_name_heb, return_p=True): 
    abs_eigvals = [np.abs(x) for x in eig_vals]
    sum_eigvals = sum(abs_eigvals)
    abs_norm_eigvals = [x/sum_eigvals for x in abs_eigvals]
    # Bokeh Scree plot
    x = list(range(len(abs_norm_eigvals)))
    y = abs_norm_eigvals
    scree_title = 'Scree plot - ' + sheet_name_heb
    p = figure(plot_width=800
, plot_height=400, title=scree_title)# x_range=len(abs_eigvals))
    # add both a line and circles on the same plot
    p.xaxis.axis_label = "Index of Eigen value"
    p.yaxis.axis_label = "Amount of variance explained"
    p.multi_line([x, x] , [y, [1/len(x) for x_ in x]], line_width=[2, 7], color=['navy', 'firebrick',],)
    p.circle(x, y, fill_color="purple", size=8, alpha=0.5)
    if return_p: 
        return p
    else: 
        show(p)
        
def calc_eig_n_plot_scree(data_f, sheet_name_heb, return_p=True): 
    eig_vals, eig_vecs = calc_eig_vals(data_f[sheet_name_heb])
    if return_p: 
        return plot_scree(eig_vals, sheet_name_heb, return_p)
    else: 
        plot_scree(eig_vals, sheet_name_heb, return_p)

def plot_scree_tabs(data, sheet_names_heb,): 
    tabs = [Panel(child=calc_eig_n_plot_scree(data, sheet_i, return_p=True), title=sheet_i) for sheet_i in sheet_names_heb]
    tabs = Tabs(tabs=list(tabs))
    show(tabs)

# Receive eigen pairs and return dataframe with sheet name and vectors
def pca_vec_row(eig_pairs_abs_, data_sheet, num_vectors=5, num_vec_components=10): 
    row = []
    for i in range(num_vectors): 
        feat_indexes = get_ind_largest_feats(eig_pairs_abs_[i][1], n=num_vec_components, return_vals=False)
        #print(feat_indexes)
        feature_names = data_sheet.columns[feat_indexes]
        #print('**********'.join(feature_names))
        row.append('**********'.join(feature_names))
    return row

def pca_vec_table(data, sheet_names_heb, num_vec_comp=10): 
    pca_trends_ = pd.DataFrame(columns=['Sheet', 'Vector_0', 'Vector_1', 'Vector_2', 'Vector_3', 'Vector_4'])
    for index, sheet_name_heb in enumerate(sheet_names_heb): 
        # Calculate list for df_i
        df_i = data[sheet_name_heb]
        # Normalizing data: Substracting mean and dividing by variance. 
        X_std = StandardScaler().fit_transform(df_i)
        # Use Numpy's covariance and calculate Eigenvectors and Eigenvalues.  
        cov_mat = np.cov(X_std.T)
        eig_vals, eig_vecs = np.linalg.eig(cov_mat)
        # eig_vals eig_vecs
        abs_eigvals = [np.abs(x) for x in eig_vals]
        sum_eigvals = sum(abs_eigvals)
        abs_norm_eigvals = [x/sum_eigvals for x in abs_eigvals]
        # Make a list of (eigenvalue, eigenvector) tuples
        eig_pairs_abs = [(np.abs(eig_vals[i]), eig_vecs[:,i]) for i in range(len(eig_vals))]
        eig_pairs_abs.sort(key=lambda x: x[0], reverse=True)
        #print(pca_vec_row(eig_pairs_abs, df_par_clean['קובץ נפות'] ))
        new_row = pca_vec_row(eig_pairs_abs, df_par_clean[sheet_name_heb], num_vec_components=num_vec_comp)
        #print(new_row)
        pca_trends_.loc[index] = [sheet_name_heb] + new_row  
    return pca_trends_

def single_sheet_pca(sheet_name_heb, processed_pca_sheet, num_elems=7, ): 
    cols = ['גורם ראשי ' + str(i) for i in range(5)]
    row_index = processed_pca_sheet[processed_pca_sheet['Sheet'] == sheet_name_heb].index.tolist()[0]
    pca_elements = [processed_pca_sheet['Vector_' + str(i) ][row_index].split('\n')[:num_elems] for i in range(5)]
    return pd.DataFrame(dict((col, pca_element) for col, pca_element in zip(cols, pca_elements)))


# TODO: 
#1) Add other params to display with hover
#2) Add heatmap bar for color of dots
#3) Add legend for size of dots
def toolTip_scatter_5d(sheet, original_sheet, x_var, y_var, size_var, color_var, hover_var, sheet_name, min_size=4, return_p=True):
    sum_size = sum(sheet[size_var])
    source = ColumnDataSource(
        data=dict(
            x = [x for x in sheet[x_var]], 
            y = [x for x in sheet[y_var]],
            size_norm =  [min_size + 2*min_size*x/sum_size for x in sheet[size_var]], 
            size_ = [x for x in sheet[size_var]],
            #color_ = [5 + 5*x for x in sheet[color_var]],
            color_ = ["#%02x%02x%02x" % (int(r), int(g), int(b)) for r, g, b, _ in 
                                255*mpl.cm.viridis(mpl.colors.Normalize()(sheet[color_var]))],
            color_hover = [x for x in sheet[color_var]],
            hover_title = [x for x in original_sheet[hover_var]][1:],)) #In original sheet first english row not removed

    hover = HoverTool(
        tooltips=[
            ("Name", "@hover_title"),
            (color_var+' (color)', '@color_hover'),
            (size_var+ ' (radius)', "@size_"),
            ("(x,y)", "($x, $y)"),])
#             ("Frequency", "@frequency"),]) # TODO: remove fo
    TOOLS = 'pan, box_zoom,box_select,resize,reset, wheel_zoom'

    p = figure(title='ייצוג דו-ממדי של '+ sheet_name,
                tools=TOOLS, plot_width=700, plot_height=400, )
    p.add_tools(hover)
    p.xaxis[0].axis_label = x_var
    if len(y_var)>30: 
        p.yaxis[0].axis_label = y_var[:30]

        #         t= y_var.split(' ')
#         t.insert(int(len(y_var.split(' '))/2), ' \n\r ')
#         p.yaxis[0].axis_label = ''.join(t)
         #' '.split(y_var).insert(int(len(' '.split(y_var))/2), '\n')
    else: 
        p.yaxis[0].axis_label = y_var
#     text_font_size = "25px"
    p.scatter('x', 'y', radius='size_norm', color='color_',
              source=source,fill_alpha=0.6,) #fill_color=colors,
    p.title.align = "center"
    p.title.text_font_size = "25px"
    if return_p: 
        return p
    else: 
        show(p)

def plot_scatter_tabs(data_dicts): 
    print('גודל דגימות: ,', data_dicts[0]['size_var'])
    print('צבע דגימות: ,', data_dicts[0]['color_var'])
    tabs = [Panel(child=toolTip_scatter_5d(**data_dict, return_p=True), title=str(ind)) for ind, data_dict in 
            zip([x for x in 'ABCDEFGHIJ'], data_dicts)]
    tabs = Tabs(tabs=list(tabs))
    show(tabs)

def create_data_dict(x_var, y_var, sheet_name_heb, sheet, ori_sheet,  size_var, color_var, hover_var): 
    return {'sheet_name': sheet_name_heb,
            'sheet': sheet[sheet_name_heb], 
             'original_sheet': ori_sheet[sheet_name_heb], 
             'x_var': x_var,
             'y_var': y_var,
             'size_var': size_var,
             'color_var': color_var,
             'hover_var': hover_var,}

def create_data_dicts(x_s, y_s, min_sizes, sheet_name_heb, df_c, df_ori, size_, color_, hover_): 
    return [create_data_dict(x, y, sheet_name_heb, df_c, df_ori,  size_, color_, hover_)
            for x, y, size_var in zip(x_s, y_s, min_sizes)]
